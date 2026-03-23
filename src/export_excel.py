"""
Export OpenAI Whisper transcriptions to Excel.
One tab per participant, just sentence number and transcription.
"""

import json
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
TRANSCRIPTIONS_DIR = BASE_DIR / "output" / "transcriptions" / "openai"
OUTPUT_DIR = BASE_DIR / "output"

AUDIO_IDS = [
    ("038010_EIT-2A", "38010-2A"),
    ("038011_EIT-1A", "38011-1A"),
    ("038012_EIT-2A", "38012-2A"),
    ("038015_EIT-1A", "38015-1A"),
]


def export():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for audio_id, sheet_name in AUDIO_IDS:
        json_path = TRANSCRIPTIONS_DIR / f"{audio_id}.json"
        with open(json_path) as f:
            data = json.load(f)

        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = "Sentence"
        ws["B1"] = "Transcription"
        for cell in [ws["A1"], ws["B1"]]:
            cell.font = openpyxl.styles.Font(bold=True)

        for i in range(30):
            row = i + 2
            ws[f"A{row}"] = i + 1
            ws[f"B{row}"] = data[i]["text"] if i < len(data) else ""

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 60

    out_path = OUTPUT_DIR / "OpenAI_Whisper_Transcriptions.xlsx"
    wb.save(str(out_path))
    print(f"Saved to {out_path}")


if __name__ == "__main__":
    export()
